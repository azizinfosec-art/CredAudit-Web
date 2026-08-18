import json
import io
import os
import subprocess
import sys
import tempfile
import unittest
from unittest import mock
from pathlib import Path
import zipfile


def run_cli(args, cwd=None):
    cmd = [sys.executable, "-m", "credaudit"] + args
    return subprocess.run(cmd, cwd=cwd, capture_output=True, text=True, check=False)


def write_file(p: Path, content: str, encoding: str = "utf-8"):
    p.write_text(content, encoding=encoding)
    return p


def write_xlsx_with_password_table(p: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"
    ws["Y1"] = "password"
    ws["Z1"] = "user "
    ws["Y2"] = "apxc@s0203"
    ws["Z2"] = "ahmed"
    ws2 = wb.create_sheet("Loose")
    ws2["B1"] = "KZMmzxw0@saa"
    ws2["C1"] = "Bader"
    wb.save(p)
    return p


def write_xlsx_with_mixed_password_headers(p: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["System", "URL", "username", "password"])
    ws.append(["prod", "https://prod.example", "alice", "Stage123!"])
    ws.append(["backup", "https://backup.example", "bob", "Backup456!"])
    wb.save(p)
    return p


def write_xlsx_with_sparse_password_row(p: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["A1"] = "password"
    ws["C1"] = "NotAdjacent123!"
    wb.save(p)
    return p


def load_json_array(p: Path):
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)


class TestCliScan(unittest.TestCase):
    def test_full_scan_default_per_file_timeout_is_two_seconds(self):
        from credaudit import cli

        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            out = tmp / "out"
            cfg = mock.Mock()
            cfg.include_ext = [".txt"]
            cfg.include_glob = []
            cfg.exclude_glob = []
            cfg.threads = 1
            cfg.workers = 1
            cfg.cache_file = str(tmp / "cache.json")
            cfg.entropy_min_length = 20
            cfg.entropy_threshold = 4.0

            with mock.patch("credaudit.cli.Config.from_yaml", return_value=cfg), \
                 mock.patch("credaudit.cli.collect_files", return_value=[str(tmp / "secrets.txt")]), \
                 mock.patch("credaudit.cli.scan_paths", return_value=([], 0)) as scan_mock, \
                 mock.patch("credaudit.cli.print_banner"), \
                 mock.patch("sys.stdout", io.StringIO()):
                rc = cli.main([
                    "scan",
                    str(tmp),
                    "-o", str(out),
                    "--full",
                    "--no-cache",
                    "--no-ndjson",
                    "--no-banner",
                ])

            self.assertEqual(rc, 0)
            self.assertEqual(scan_mock.call_args.kwargs.get("per_file_timeout"), 2.0)

    def test_scan_ndjson_and_json_roundtrip(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            # Create a guaranteed hit
            write_file(tmp / "secrets.txt", "password: Abcd1234\n")
            out_dir = tmp / "out"
            nd = out_dir / "findings.ndjson"
            # --no-timestamp keeps report file deterministic (report.json)
            res = run_cli([
                "scan", "-p", str(tmp), "-o", str(out_dir), "--no-cache",
                "--ndjson-out", str(nd),
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            self.assertTrue(nd.exists() and nd.stat().st_size > 0)
            # Check first NDJSON line has required fields
            first_line = nd.read_text(encoding="utf-8").splitlines()[0]
            obj = json.loads(first_line)
            for k in ["ts", "file", "rule", "severity", "redacted", "context", "line"]:
                self.assertIn(k, obj)
            # Check JSON report
            j = out_dir / "report.json"
            self.assertTrue(j.exists(), "report.json not found")
            arr = load_json_array(j)
            self.assertTrue(any(f.get("rule") == "PasswordValueAssignment" for f in arr))

    def test_examples_command_prints_copy_paste_commands(self):
        res = run_cli(["examples", "--no-banner"])
        self.assertEqual(res.returncode, 0, res.stderr)
        self.assertIn("CredAudit example commands", res.stdout)
        self.assertIn("credaudit ./client-data", res.stdout)
        self.assertIn("credaudit scan -p . --full --safe --formats sarif", res.stdout)

    def test_utf16_text_file_detects_password_assignment(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            secret = write_file(tmp / "utf16.txt", "password: Secret123!\n", encoding="utf-16")
            out = tmp / "out"
            res = run_cli([
                str(secret),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertTrue(
                any(f.get("rule") == "PasswordValueAssignment" and f.get("match") == "Secret123!" for f in arr),
                arr,
            )

    def test_only_rules_filters_findings(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\napi_key=sk-abcde1234567890\n")
            out = tmp / "out"
            res = run_cli([
                "scan", "-p", str(tmp), "-o", str(out), "--no-cache",
                "--formats", "json",
                "--no-timestamp",
                "--only-rules", "PasswordAssignment"
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertTrue(arr, "no findings produced")
            self.assertTrue(all(f.get("rule") == "PasswordAssignment" for f in arr))

    def test_cache_is_invalidated_when_only_rules_change(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\napi_key=sk-abcde1234567890\n")
            out = tmp / "out"
            cache = tmp / "cache.json"

            first = run_cli([
                "scan", "-p", str(tmp), "-o", str(out),
                "--cache-file", str(cache),
                "--formats", "json",
                "--no-timestamp",
                "--full",
                "--raw",
                "--no-ndjson",
                "--only-rules", "PasswordAssignment",
            ])
            self.assertEqual(first.returncode, 0, first.stderr)
            first_arr = load_json_array(out / "report.json")
            self.assertTrue(any(f.get("rule") == "PasswordAssignment" for f in first_arr), first_arr)
            self.assertFalse(any(f.get("rule") == "APIKeyGeneric" for f in first_arr), first_arr)

            second = run_cli([
                "scan", "-p", str(tmp), "-o", str(out),
                "--cache-file", str(cache),
                "--formats", "json",
                "--no-timestamp",
                "--full",
                "--raw",
                "--no-ndjson",
                "--only-rules", "APIKeyGeneric",
            ])
            self.assertEqual(second.returncode, 0, second.stderr)
            second_arr = load_json_array(out / "report.json")
            self.assertTrue(any(f.get("rule") == "APIKeyGeneric" for f in second_arr), second_arr)
            self.assertFalse(any(f.get("rule") == "PasswordAssignment" for f in second_arr), second_arr)

    def test_safe_shortcut_redacts_json_and_skips_cache(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\n")
            out = tmp / "out"
            cache = tmp / "cache.json"
            res = run_cli([
                "scan",
                str(tmp),
                "-o", str(out),
                "--cache-file", str(cache),
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            report = out / "report.json"
            self.assertTrue(report.exists(), "safe shortcut did not create report.json")
            text = report.read_text(encoding="utf-8")
            self.assertNotIn("Abcd1234", text)
            self.assertIn("A****4", text)
            self.assertFalse(cache.exists(), "safe mode should not write a raw findings cache")

    def test_safe_shortcut_fast_defaults_to_small_txt_only(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            small_txt = write_file(tmp / "small.txt", "password: Abcd1234\n")
            write_file(tmp / "secret.json", "{\"password\":\"Json1234\"}\n")
            write_file(tmp / "large.txt", ("A" * 11000) + "\npassword: Large1234\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertTrue(arr, "small .txt file should be scanned")
            files = {Path(f.get("file", "")).name for f in arr}
            self.assertIn(small_txt.name, files)
            self.assertNotIn("secret.json", files)
            self.assertNotIn("large.txt", files)

    def test_scan_command_defaults_to_safe_fast(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            small_txt = write_file(tmp / "small.txt", "password: Abcd1234\n")
            write_file(tmp / "secret.env", "password=Env1234\n")
            write_file(tmp / "large.txt", ("A" * 11000) + "\npassword: Large1234\n")
            out = tmp / "out"
            cache = tmp / "cache.json"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--cache-file", str(cache),
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            text = (out / "report.json").read_text(encoding="utf-8")
            self.assertNotIn("Abcd1234", text)
            self.assertIn("A****4", text)
            self.assertFalse(cache.exists(), "default safe mode should not write a raw findings cache")
            arr = load_json_array(out / "report.json")
            files = {Path(f.get("file", "")).name for f in arr}
            self.assertIn(small_txt.name, files)
            self.assertNotIn("secret.env", files)
            self.assertNotIn("large.txt", files)

    def test_full_raw_opt_out_uses_standard_scope_and_cache(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\n")
            write_file(tmp / "secret.env", "password=Env1234\n")
            out = tmp / "out"
            cache = tmp / "cache.json"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--cache-file", str(cache),
                "--formats", "json",
                "--no-timestamp",
                "--full",
                "--raw",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            text = (out / "report.json").read_text(encoding="utf-8")
            self.assertIn("Abcd1234", text)
            self.assertIn("Env1234", text)
            self.assertTrue(cache.exists(), "raw standard mode should write the findings cache")
            arr = load_json_array(out / "report.json")
            files = {Path(f.get("file", "")).name for f in arr}
            self.assertIn("secret.env", files)

    def test_include_glob_allows_explicit_non_default_text_extension(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secret.yaml", "password: Yaml1234\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--include-glob", "**/*.yaml",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            files = {Path(f.get("file", "")).name for f in arr}
            self.assertIn("secret.yaml", files)

    def test_explicit_xlsx_file_scans_password_table_by_default(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            workbook = write_xlsx_with_password_table(tmp / "MyFile.xlsx")
            out = tmp / "out"
            res = run_cli([
                str(workbook),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertTrue(
                any(
                    f.get("rule") == "PasswordValueAssignment" and f.get("match") == "apxc@s0203"
                    for f in arr
                ),
                arr,
            )
            self.assertFalse(
                any(f.get("rule") in {"PasswordAssignment", "PasswordValueAssignment"} and f.get("match") == "user" for f in arr),
                arr,
            )

    def test_xlsx_mixed_headers_scan_multiple_password_rows(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            workbook = write_xlsx_with_mixed_password_headers(tmp / "mixed.xlsx")
            out = tmp / "out"
            res = run_cli([
                str(workbook),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            password_matches = {
                f.get("match")
                for f in arr
                if f.get("rule") == "PasswordValueAssignment"
            }
            self.assertIn("Stage123!", password_matches)
            self.assertIn("Backup456!", password_matches)

    def test_xlsx_sparse_row_does_not_pair_non_adjacent_password_value(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            workbook = write_xlsx_with_sparse_password_row(tmp / "sparse.xlsx")
            out = tmp / "out"
            res = run_cli([
                str(workbook),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertFalse(
                any(
                    f.get("rule") in {"PasswordAssignment", "PasswordValueAssignment"} and f.get("match") == "NotAdjacent123!"
                    for f in arr
                ),
                arr,
            )

    def test_overlapping_rules_are_deduplicated_by_secret_value(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Secret123!;\napi_key=sk-abcde1234567890,\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--raw",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            findings_by_line = {}
            for finding in arr:
                findings_by_line.setdefault(finding.get("line"), []).append(finding)
            self.assertEqual([f.get("rule") for f in findings_by_line.get(1, [])], ["PasswordValueAssignment"])
            self.assertEqual([f.get("match") for f in findings_by_line.get(1, [])], ["Secret123!"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(2, [])], ["APIKeyGeneric"])
            self.assertEqual([f.get("match") for f in findings_by_line.get(2, [])], ["sk-abcde1234567890"])

    def test_username_and_password_keyword_indicators_are_reported(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "indicators.txt", "username=admin\nplease rotate password soon\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            rules_by_line = {}
            for finding in arr:
                rules_by_line.setdefault(finding.get("line"), []).append(finding.get("rule"))
            self.assertEqual(rules_by_line.get(1), ["UsernameAssignment"])
            self.assertEqual(rules_by_line.get(2), ["PasswordKeyword"])

    def test_standalone_password_candidates_are_reported_by_default(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(
                tmp / "candidates.txt",
                "myo@193\nmISX%%13402\npassword: myo@193\njohn@example.com\npython-docx\nCredAudit v{VERSION}\n",
            )
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--raw",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            findings_by_line = {}
            for finding in arr:
                findings_by_line.setdefault(finding.get("line"), []).append(finding)
            self.assertEqual([f.get("rule") for f in findings_by_line.get(1, [])], ["PasswordCandidate"])
            self.assertEqual([f.get("match") for f in findings_by_line.get(1, [])], ["myo@193"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(2, [])], ["PasswordCandidate"])
            self.assertEqual([f.get("match") for f in findings_by_line.get(2, [])], ["mISX%%13402"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(3, [])], ["PasswordValueAssignment"])
            self.assertNotIn(4, findings_by_line)
            self.assertNotIn(5, findings_by_line)
            self.assertNotIn(6, findings_by_line)

    def test_username_on_line_before_password_is_reported(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(
                tmp / "paired.txt",
                "admin\nmyo@193\nusername=alice\npassword: Secret123!\nthis is just text\nmISX%%13402\njohn@example.com\nmISX%%13402\n",
            )
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--raw",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            findings_by_line = {}
            for finding in arr:
                findings_by_line.setdefault(finding.get("line"), []).append(finding)
            self.assertEqual([f.get("rule") for f in findings_by_line.get(1, [])], ["UsernameNearPassword"])
            self.assertEqual([f.get("severity") for f in findings_by_line.get(1, [])], ["High"])
            self.assertEqual([f.get("match") for f in findings_by_line.get(1, [])], ["admin"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(2, [])], ["PasswordCandidate"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(3, [])], ["UsernameNearPassword"])
            self.assertEqual([f.get("severity") for f in findings_by_line.get(3, [])], ["High"])
            self.assertEqual([f.get("match") for f in findings_by_line.get(3, [])], ["alice"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(4, [])], ["PasswordValueAssignment"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(6, [])], ["PasswordCandidate"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(7, [])], ["UsernameNearPassword"])
            self.assertEqual([f.get("severity") for f in findings_by_line.get(7, [])], ["High"])
            self.assertEqual([f.get("match") for f in findings_by_line.get(7, [])], ["john@example.com"])
            self.assertEqual([f.get("rule") for f in findings_by_line.get(8, [])], ["PasswordCandidate"])

    def test_same_line_credential_pairs_in_small_txt_files(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(
                tmp / "pairs.txt",
                "\n".join([
                    "admin:Secret123!",
                    "alice@example.com:password",
                    "username: alice",
                    "Content-Type:text/html",
                    "module:ClassName",
                    "https://example.com/login",
                ]) + "\n",
            )
            write_file(tmp / "large.txt", "oversize:Secret123!\n" + ("A" * (5 * 1024 * 1024)))
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--include-ext", ".txt",
                "--max-size", "5",
                "--only-rules", "CredentialPair",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            findings_by_line = {f.get("line"): f for f in arr}
            self.assertEqual({f.get("rule") for f in arr}, {"CredentialPair"})
            self.assertEqual(findings_by_line.get(1, {}).get("match"), "Secret123!")
            self.assertEqual(findings_by_line.get(2, {}).get("match"), "password")
            self.assertNotIn(3, findings_by_line)
            self.assertNotIn(4, findings_by_line)
            self.assertNotIn(5, findings_by_line)
            self.assertNotIn(6, findings_by_line)
            files = {Path(f.get("file", "")).name for f in arr}
            self.assertEqual(files, {"pairs.txt"})

    def test_min_confidence_filters_and_exports_evidence(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(
                tmp / "client-creds.txt",
                "\n".join([
                    "admin:Secret123!",
                    "password: Abcd1234",
                    "mISX%%13402",
                    "api_key=sk-abcde1234567890",
                    "please rotate password soon",
                ]) + "\n",
            )
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--include-ext", ".txt",
                "--max-size", "5",
                "--min-confidence", "80",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertTrue(arr)
            self.assertTrue(all(int(f.get("confidence", 0)) >= 80 for f in arr), arr)
            rules = {f.get("rule") for f in arr}
            self.assertIn("CredentialPair", rules)
            self.assertIn("PasswordValueAssignment", rules)
            self.assertNotIn("PasswordCandidate", rules)
            self.assertNotIn("PasswordKeyword", rules)
            credential = next(f for f in arr if f.get("rule") == "CredentialPair")
            self.assertEqual(credential.get("finding_class"), "likely")
            self.assertEqual(credential.get("validity"), "not_applicable")
            self.assertTrue(credential.get("evidence"))
            self.assertTrue(any("username:password" in x for x in credential.get("evidence", [])))

    def test_high_confidence_console_can_show_evidence(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "client-creds.txt", "admin:Secret123!\nmISX%%13402\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--include-ext", ".txt",
                "--max-size", "5",
                "--high-confidence",
                "--show-evidence",
                "--no-cache",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            self.assertIn("Score", res.stdout)
            self.assertIn("Evidence:", res.stdout)
            self.assertIn("CredentialPair", res.stdout)
            self.assertNotIn("PasswordCandidate", res.stdout)

    def test_passwords_intent_finds_password_shapes_without_rule_names(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(
                tmp / "client-creds.txt",
                "\n".join([
                    "admin:Secret123!",
                    "password: Abcd1234",
                    "mISX%%13402",
                    "please rotate password soon",
                ]) + "\n",
            )
            write_file(tmp / "config.json", "{\"password\":\"Json1234\"}\n")
            out = tmp / "out"
            res = run_cli([
                "passwords",
                str(tmp),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            rules = {f.get("rule") for f in arr}
            self.assertIn("CredentialPair", rules)
            self.assertIn("PasswordValueAssignment", rules)
            self.assertIn("PasswordCandidate", rules)
            self.assertNotIn("PasswordKeyword", rules)
            self.assertFalse(any(f.get("match") == "sk-abcde1234567890" for f in arr), arr)
            self.assertEqual({Path(f.get("file", "")).name for f in arr}, {"client-creds.txt"})
            self.assertTrue(all(int(f.get("confidence", 0)) >= 50 for f in arr), arr)

    def test_password_keyword_values_take_priority_and_high_confidence(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(
                tmp / "passwords.txt",
                "password: Secret123!\npasswd=Root123!\npasscode: Code1234\napi_key=sk-abcde1234567890\n",
            )
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--include-ext", ".txt",
                "--max-size", "5",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            by_match = {f.get("match"): f for f in arr}
            for value in ["Secret123!", "Root123!", "Code1234"]:
                self.assertEqual(by_match.get(value, {}).get("rule"), "PasswordValueAssignment")
                self.assertGreaterEqual(int(by_match.get(value, {}).get("confidence", 0)), 95)
                self.assertEqual(by_match.get(value, {}).get("severity"), "Critical")
                self.assertTrue(
                    any("after a password/pass/pwd keyword" in x for x in by_match.get(value, {}).get("evidence", [])),
                    by_match.get(value),
                )
            self.assertEqual(by_match.get("sk-abcde1234567890", {}).get("rule"), "APIKeyGeneric")

    def test_metadata_style_colon_pairs_are_not_scored_like_passwords(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(
                tmp / "headers.txt",
                "Content-Transfer-Encoding: base64\npassword: Secret123!\n",
            )
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--raw",
                "--no-cache",
                "--include-ext", ".txt",
                "--max-size", "5",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            by_match = {f.get("match"): f for f in arr}
            metadata = by_match.get("base64")
            self.assertIsNotNone(metadata, arr)
            self.assertEqual(metadata.get("rule"), "CredentialPair")
            self.assertEqual(metadata.get("severity"), "Medium")
            self.assertLessEqual(int(metadata.get("confidence", 0)), 69)
            self.assertGreaterEqual(int(metadata.get("confidence", 0)), 60)
            self.assertTrue(
                any("metadata label" in x for x in metadata.get("evidence", [])),
                metadata,
            )
            password = by_match.get("Secret123!")
            self.assertEqual(password.get("rule"), "PasswordValueAssignment")
            self.assertEqual(password.get("severity"), "Critical")
            self.assertGreaterEqual(int(password.get("confidence", 0)), 95)

    def test_fail_on_critical_uses_confidence_based_severity(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "passwords.txt", "password: Secret123!\nContent-Transfer-Encoding: base64\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--include-ext", ".txt",
                "--max-size", "5",
                "--fail-on", "Critical",
                "--no-cache",
                "--no-banner",
            ])
            self.assertEqual(res.returncode, 2, res.stdout + res.stderr)
            self.assertIn("Critical", res.stdout)

    def test_precommit_blocks_critical_findings(self):
        repo = Path(__file__).resolve().parents[2]
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            secret = write_file(tmp / "secrets.txt", "password: Secret123!\n")
            res = subprocess.run(
                [sys.executable, "scripts/precommit_scan.py", str(secret)],
                cwd=repo,
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(res.returncode, 2, res.stdout + res.stderr)
            self.assertIn("[Critical]", res.stdout)

    def test_shortcut_without_formats_prints_redacted_console(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            self.assertIn("Findings (redacted)", res.stdout)
            self.assertIn("Severity Rule                     Line   File", res.stdout)
            self.assertIn("PasswordValueAssignment", res.stdout)
            self.assertIn("A****4", res.stdout)
            self.assertNotIn("Abcd1234", res.stdout)
            self.assertFalse((out / "report.json").exists(), "console mode should not write report.json")
            nd = out / "findings.ndjson"
            self.assertTrue(nd.exists() and nd.stat().st_size > 0, "default NDJSON stream was not written")
            self.assertIn("NDJSON:", res.stdout)
            first = json.loads(nd.read_text(encoding="utf-8").splitlines()[0])
            self.assertEqual(first.get("rule"), "PasswordValueAssignment")
            self.assertIn("A****4", first.get("redacted", ""))
            self.assertNotIn("Abcd1234", nd.read_text(encoding="utf-8"))

    def test_no_ndjson_disables_default_stream(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--no-ndjson",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            self.assertIn("Findings (redacted)", res.stdout)
            self.assertFalse((out / "findings.ndjson").exists())

    def test_formats_print_clickable_file_urls(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\n")
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--formats", "html", "json",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            html_reports = list(out.glob("report_*.html"))
            json_reports = list(out.glob("report_*.json"))
            ndjson_reports = list(out.glob("findings_*.ndjson"))
            self.assertEqual(len(html_reports), 1)
            self.assertEqual(len(json_reports), 1)
            self.assertEqual(len(ndjson_reports), 1)
            html = html_reports[0].resolve().as_uri()
            json_report = json_reports[0].resolve().as_uri()
            self.assertIn("Report URLs:", res.stdout)
            self.assertIn(f"HTML: {html}", res.stdout)
            self.assertIn(f"JSON: {json_report}", res.stdout)
            self.assertIn("NDJSON:", res.stdout)

    def test_html_generated_with_template(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "secrets.txt", "password: Abcd1234\n")
            out = tmp / "out"
            res = run_cli([
                "scan", "-p", str(tmp), "-o", str(out), "--no-cache",
                "--formats", "html", "--timestamp"
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            # Find any HTML report in out dir
            htmls = list(out.glob("report_*.html")) or list(out.glob("report.html"))
            self.assertTrue(htmls, "no HTML report produced")
            html = htmls[0].read_text(encoding="utf-8", errors="ignore")
            # Sanity checks for the new chrome
            self.assertIn("CredAudit Report", html)
            self.assertTrue("id=\"tbl\"" in html or "<table id=\"tbl\"" in html)

    def test_scan_har_responses(self):
        """Ensure HAR response body scanning finds secrets in JSON responses (quoted style)."""
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            har = tmp / "traffic.har"
            har_obj = {
                "log": {
                    "version": "1.2",
                    "creator": {"name": "test", "version": "1.0"},
                    "entries": [
                        {
                            "request": {"method": "GET", "url": "https://example.local/"},
                            "response": {
                                "status": 200,
                                "content": {
                                    "mimeType": "application/json",
                                    "text": "{\"password\":\"Abcd1234\"}"
                                }
                            }
                        }
                    ]
                }
            }
            har.write_text(json.dumps(har_obj), encoding="utf-8")
            out = tmp / "out"
            res = run_cli([
                "scan", "-p", str(har), "-o", str(out), "--no-cache",
                "--include-ext", ".har",
                "--har-include", "responses",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            ok = any((f.get("rule") in ("PasswordAssignment","PasswordAssignmentLoose","PasswordValueAssignment","PasswordValueAssignmentLoose")) for f in arr)
            self.assertTrue(ok, f"No password-like finding in HAR: {arr}")

    def test_scan_archive_zip(self):
        """Ensure ZIP archives are expanded and findings remap to 'zip!inner' paths."""
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            zpath = tmp / "a.zip"
            # create zip with secrets.txt
            with zipfile.ZipFile(zpath, "w") as z:
                inner_name = "secrets.txt"
                z.writestr(inner_name, "password: Abcd1234\n")
            out = tmp / "out"
            res = run_cli([
                "scan", "-p", str(zpath), "-o", str(out), "--no-cache",
                "--scan-archives", "--archive-depth", "1",
                "--include-ext", ".zip",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertTrue(arr)
            # Check alias path like a.zip!secrets.txt
            self.assertTrue(any(".zip!" in (f.get("file") or "") for f in arr))

    def test_scan_archive_zip_har_member(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            zpath = tmp / "traffic.zip"
            har_obj = {
                "log": {
                    "version": "1.2",
                    "creator": {"name": "test", "version": "1.0"},
                    "entries": [
                        {
                            "request": {"method": "GET", "url": "https://example.local/"},
                            "response": {
                                "status": 200,
                                "content": {
                                    "mimeType": "application/json",
                                    "text": "{\"password\":\"Abcd1234\"}"
                                }
                            }
                        }
                    ]
                }
            }
            with zipfile.ZipFile(zpath, "w") as z:
                z.writestr("traffic.har", json.dumps(har_obj))
            out = tmp / "out"
            res = run_cli([
                "scan", "-p", str(zpath), "-o", str(out), "--no-cache",
                "--scan-archives", "--archive-depth", "1",
                "--include-ext", ".zip",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertTrue(
                any(f.get("rule") in {"PasswordAssignment", "PasswordValueAssignment"} for f in arr),
                arr,
            )

    def test_archive_member_respects_max_size_limit(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            zpath = tmp / "large.zip"
            with zipfile.ZipFile(zpath, "w", compression=zipfile.ZIP_DEFLATED) as z:
                z.writestr("large.txt", "password: Secret123!\n" + ("A" * 5000))
            out = tmp / "out"
            res = run_cli([
                "scan", "-p", str(zpath), "-o", str(out), "--no-cache",
                "--scan-archives", "--archive-depth", "1",
                "--include-ext", ".zip",
                "--max-size-kb", "1",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertEqual(arr, [])

    def test_config_can_disable_entropy_rule(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            write_file(tmp / "token.txt", "abcdefghijklmnopqrstuvwxyzABCDEFGH\n")
            cfg = tmp / "config.yaml"
            cfg.write_text(
                "include_ext: ['.txt']\n"
                "rules:\n"
                "  enable_entropy: false\n",
                encoding="utf-8",
            )
            out = tmp / "out"
            res = run_cli([
                str(tmp),
                "-o", str(out),
                "--config", str(cfg),
                "--raw",
                "--formats", "json",
                "--no-timestamp",
            ])
            self.assertEqual(res.returncode, 0, res.stderr)
            arr = load_json_array(out / "report.json")
            self.assertFalse(any(f.get("rule") == "HighEntropyString" for f in arr), arr)


if __name__ == "__main__":
    unittest.main()
